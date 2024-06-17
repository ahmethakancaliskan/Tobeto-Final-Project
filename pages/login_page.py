import time
from selenium.webdriver.common.by import By
from pages.base_page import BasePage
import openpyxl


class LoginPage(BasePage):
    __mailInput = (By.NAME, "email")
    __passwordInput = (By.NAME, "password")
    __loginButton = (By.CSS_SELECTOR, "button[type='submit']")
    __warning_massage = (By.CSS_SELECTOR, ".toast-body")
    __empty_inputs = (By.CSS_SELECTOR, "[action='#'] p")

    def enter_email(self, mail):
        self.scroll_page_down(250)
        self.find_element(self.__mailInput).send_keys(mail)
        time.sleep(5)

    def enter_password(self, password):
        self.wait_until_element_visible(self.__passwordInput)
        self.find_element(self.__passwordInput).send_keys(password)
        time.sleep(5)

    def click_login_button(self):
        self.wait_until_element_visible(self.__loginButton)
        self.find_element(self.__loginButton).click()
        time.sleep(1)

    def get_login_warning_massage(self):
        self.wait_until_element_visible(self.__warning_massage)
        massage = self.find_element(self.__warning_massage)
        return massage.text

    def get_empty_inputs_error_massage(self):
        massage = self.find_element(self.__empty_inputs)
        return massage.text

    @staticmethod
    def get_data():
        excelFile = openpyxl.load_workbook("../data/invalidLogin.xlsx")  # dosya açılıyor
        sheet = excelFile["Sheet1"]  # hangi sayfada çalıştığımızı belirtiyoruz
        rows = sheet.max_row  # kaçıncı satıra kadar veri var?
        data = []
        for i in range(2, rows + 1):
            username = sheet.cell(i, 1).value  # i. satırın 1. hücresi(cell) username'i tutuyor
            password = sheet.cell(i, 2).value
            data.append((username, password))
        return data
